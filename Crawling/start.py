

# 엑셀파일
from openpyxl import load_workbook
from selenium.common.exceptions import InvalidArgumentException, WebDriverException

from Crawling.ani_detail_crawl import namuwiki_crawl, google_crawl, cueing_crawl

print("\n 1 : 나무위키 크롤링\n")
print(" 2 : 구글 크롤링\n")
print(" 3 : 츄잉 크롤링\n")
select = input(" 선택 : ")

# 나무위키 크롤링
if select == "1":
    xlsx_list = ["ani_a.xlsx", "ani_b.xlsx", "ani_c.xlsx", "ani_d.xlsx", "ani_e.xlsx", "ani_f.xlsx", "ani_g.xlsx",
                 "ani_h.xlsx", "ani_i.xlsx", "ani_j.xlsx", "ani_k.xlsx", "ani_l.xlsx", "ani_m.xlsx", "ani_n.xlsx",
                 "ani_o.xlsx", "ani_p.xlsx"]

    for xl in xlsx_list:
        work_book = load_workbook(xl)
        sheet = work_book['Sheet']
        for no in range(2, len(sheet["A"]) + 1):
            try:
                namuwiki_crawl(work_book, xl, sheet["H"+str(no)].value, sheet["B"+str(no)])
            except InvalidArgumentException:
                print("InvalidArgumentException2")

# 구글 크롤링
if select == "2":
    xlsx_list = ["ani_a.xlsx", "ani_b.xlsx", "ani_c.xlsx", "ani_d.xlsx", "ani_e.xlsx", "ani_f.xlsx", "ani_g.xlsx",
                 "ani_h.xlsx", "ani_i.xlsx", "ani_j.xlsx", "ani_k.xlsx", "ani_l.xlsx", "ani_m.xlsx", "ani_n.xlsx",
                 "ani_o.xlsx", "ani_p.xlsx"]

    for xl in xlsx_list:
        work_book = load_workbook(xl)
        sheet = work_book['Sheet']
        for no in range(2, len(sheet["A"]) + 1):
            try:
                google_crawl(work_book, xl)
            except InvalidArgumentException:
                print("InvalidArgumentException2")


# 츄잉 크롤링
if select == "3":
    xlsx_list = ["ani_a.xlsx", "ani_b.xlsx", "ani_c.xlsx", "ani_d.xlsx", "ani_e.xlsx", "ani_f.xlsx", "ani_g.xlsx",
                 "ani_h.xlsx", "ani_i.xlsx", "ani_j.xlsx", "ani_k.xlsx", "ani_l.xlsx", "ani_m.xlsx", "ani_n.xlsx",
                 "ani_o.xlsx", "ani_p.xlsx"]

    for xl in xlsx_list:
        work_book = load_workbook(xl)
        sheet = work_book['Sheet']
        for no in range(2, len(sheet["A"]) + 1):
            try:
                cueing_crawl(work_book, xl)
            except InvalidArgumentException:
                print("InvalidArgumentException2")
