import datetime
import os
import time
from openpyxl import load_workbook
from openpyxl.styles import Font

update_xl = "ani_update.xlsx"
update_book = load_workbook(update_xl)
update_sheet = update_book["Sheet"]

xl = "ani_detail.xlsx"
work_book = load_workbook(xl)
sheet = work_book["Sheet"]

last_index = len(update_sheet["A"]) + len(sheet["A"])

no = 2
for kk in range(len(sheet["A"]) + 1, last_index):
    sheet["A" + str(kk)].value = update_sheet["A" + str(no)].value
    sheet["A" + str(kk)].font = Font(name="나눔고딕", color="000000")

    sheet["B" + str(kk)].value = update_sheet["B" + str(no)].value
    sheet["B" + str(kk)].font = Font(name="나눔고딕", color="000000")

    sheet["C" + str(kk)].value = update_sheet["C" + str(no)].value
    sheet["C" + str(kk)].font = Font(name="나눔고딕", color="000000")

    sheet["D" + str(kk)].value = update_sheet["D" + str(no)].value
    sheet["D" + str(kk)].font = Font(name="나눔고딕", color="000000")

    sheet["E" + str(kk)].value = update_sheet["E" + str(no)].value
    sheet["E" + str(kk)].font = Font(name="나눔고딕", color="000000")

    sheet["H" + str(kk)].value = update_sheet["H" + str(no)].value
    sheet["H" + str(kk)].font = Font(name="나눔고딕", color="000000")
    no += 1

work_book.save(xl)

# 업데이트 엑셀 파일명 변경
os.rename(update_xl, str(datetime.datetime.now().date()) + "_" + update_xl)
print("\n ★ 업데이트 완료 ★")

