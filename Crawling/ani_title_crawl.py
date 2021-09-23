from openpyxl.styles import Font
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
import pandas as pd

from Settings.selenium_setting import *

# 크롤링 페이지
driver.get("https://namu.wiki/w/%EC%9D%BC%EB%B3%B8%20%EC%95%A0%EB%8B%88%EB%A9%94%EC%9D%B4%EC%85%98/%EB%AA%A9%EB%A1%9D/%E3%84%B1")

ani_data = {}
ani_title_list = []
xlsx_name = "ani_a"

# 상세 페이지 크롤링
ani_detail_list_xpath_base = "//*[@id='app']/div/div[2]/article/div[3]/div[2]/div/div/div/ul/li/div"
# ani_detail_list_xpath_base = "//*[@id='app']/div/div[2]/article/div[3]/div[2]/div/div/div/ul/li/div/a[1]"
ani_detail_title = driver.find_elements_by_xpath(ani_detail_list_xpath_base + "/a")

for index, kk in enumerate(ani_detail_title):
    ani_data[kk.text] = kk.get_attribute("href")
    ani_title_list.append(kk.text)
    # print(kk.text)

# time.sleep(2)
print(ani_data)

driver.quit()

# 엑셀파일 쓰기
work_book = Workbook()

# 시트 생성
# work_book.create_sheet('애니메이션')

# 시트 입력
sheet = work_book.active
sheet.column_dimensions["A"].width = 45

for kk in range(1, len(ani_data)):
    sheet['A' + str(kk)] = ani_title_list[kk - 1]
    sheet['A' + str(kk)].font = Font(name="나눔고딕", color="000000")
    sheet['B' + str(kk)] = ani_data[ani_title_list[kk - 1]]
    # sheet['B' + str(kk)].value = '=HYPERLINK("{}", "{}")'.format(ani_data[ani_title_list[kk - 1]], "링크")
    sheet['B' + str(kk)].font = Font(name="나눔고딕", color="50BCDF")

work_book.save(xlsx_name + ".xlsx")
