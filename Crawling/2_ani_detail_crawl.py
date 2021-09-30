

import datetime
import urllib.request

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.common.exceptions import StaleElementReferenceException, InvalidArgumentException, TimeoutException, \
    NoSuchElementException, WebDriverException
import time

chd = 'C:/dev_files/chd/chd.exe'
options = webdriver.ChromeOptions()
options.add_argument("headless")
options.add_argument("window-size=1920x1080")
options.add_argument("disable-gpu")
options.add_argument("User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, "
                     "like Gecko) Chrome/93.0.4577.82")
options.add_argument("lang=ko_KR")
driver = webdriver.Chrome(chd, options=options)

# 엑셀파일 생성
save_book = Workbook()
# 시트 입력
save_sheet = save_book.active
save_sheet.column_dimensions["A"].width = 55
save_sheet.column_dimensions["B"].width = 55
save_sheet.column_dimensions["C"].width = 55
save_sheet.column_dimensions["D"].width = 30
save_sheet.column_dimensions["E"].width = 30
save_sheet.column_dimensions["H"].width = 30

save_sheet['A1'] = "제목"
save_sheet['A1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['A1'].alignment = Alignment(horizontal='center')
save_sheet['B1'] = "장르"
save_sheet['B1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['B1'].alignment = Alignment(horizontal='center')
save_sheet['C1'] = "태그"
save_sheet['C1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['C1'].alignment = Alignment(horizontal='center')
save_sheet['D1'] = "방영일"
save_sheet['D1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['D1'].alignment = Alignment(horizontal='center')
save_sheet['E1'] = "대표이미지 파일명"
save_sheet['E1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['E1'].alignment = Alignment(horizontal='center')
save_sheet['H1'] = "링크"
save_sheet['H1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['H1'].alignment = Alignment(horizontal='center')

# 리스트 불러오기
load_book = load_workbook("anime_detail_link.xlsx")
load_sheet = load_book['Sheet']

title_xpath = "/html/body/div/div/div/article/div/h1"
genre_check_xpath = "/html/body/div/div/div/article/div/div/p/span[1]"
image_xpath = "/html/body/div/div/div/article/div/div/div/a/img"

for no in range(2, len(load_sheet["A"])):
    try:
        # 크롤링 주소 가져오기
        driver.get(load_sheet["B"+str(no)].value)

        time.sleep(3)

        # 링크 시트 삽입
        save_sheet["H"+str(no)].value = load_sheet["B"+str(no)].value

        # XPATH 찾기
        title_tag = driver.find_elements_by_xpath(title_xpath)
        genre_check_tag = driver.find_elements_by_xpath(genre_check_xpath)
        image_tag = driver.find_elements_by_xpath(image_xpath)

        # 제목 크롤링
        for title in title_tag:
            save_sheet["A"+str(no)].value = title.text

        # 장르, 방영일 크롤링
        genre_count = 1
        date_count = 1
        for aa in genre_check_tag:
            if aa.text == "장르":
                genre_xpath = "/html/body/div/div/div/article/div/div/p[" + str(genre_count) + "]/span[2]"
                genre_tag = driver.find_elements_by_xpath(genre_xpath)
                for genre in genre_tag:
                    save_sheet["B" + str(no)].value = genre.text

            if aa.text == "방영일":
                date_xpath = "/html/body/div/div/div/article/div/div/p[" + str(date_count) + "]/span[2]"
                date_tag = driver.find_elements_by_xpath(date_xpath)
                for date in date_tag:
                    save_sheet["D" + str(no)].value = date.text

            genre_count += 1
            date_count += 1

        # 대표 이미지 크롤링 및 다운로드
        for image in image_tag:
            # 시간으로 파일명 지정
            now_time = datetime.datetime.now()
            front_time = str(now_time.year) + str(now_time.month) + str(now_time.day)
            back_time = str(now_time.hour) + str(now_time.minute) + str(now_time.second) + str(now_time.microsecond)
            image_file_name = front_time + back_time + ".jpg"
            print(image_file_name)

            # 다운받을 이미지 링크
            url = image.get_attribute("src")

            # 이미지 다운로드
            urllib.request.urlretrieve(url, "../static/image/" + image_file_name)
            time.sleep(5)

            save_sheet["E" + str(no)].value = image_file_name

    except StaleElementReferenceException:
        print("StaleElementReferenceException")
    except InvalidArgumentException:
        print("InvalidArgumentException")
    except NoSuchElementException:
        print("NoSuchElementException")
    except TimeoutException:
        print("TimeoutException")
        driver.quit()

        time.sleep(5)
        options = webdriver.ChromeOptions()
        options.add_argument("headless")
        options.add_argument("window-size=1920x1080")
        options.add_argument("disable-gpu")
        options.add_argument("User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, "
                             "like Gecko) Chrome/93.0.4577.82")
        options.add_argument("lang=ko_KR")
        driver = webdriver.Chrome(chd, options=options)

    except WebDriverException:
        print("WebDriverException")
    except IndexError:
        print("IndexError")
    except:
        print("Error 발생!")

driver.quit()

save_book.save("anime_detail.xlsx")

