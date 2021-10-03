import os
import time
import datetime
import urllib.request

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from selenium.common.exceptions import NoSuchElementException, TimeoutException, StaleElementReferenceException, \
    InvalidArgumentException, WebDriverException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver

# 셀레니움 셋팅(Selenium Setting)
chd = 'C:/dev_files/chd/chd.exe'
options = webdriver.ChromeOptions()
options.add_argument("headless")
options.add_argument("window-size=1920x1080")
options.add_argument("disable-gpu")
options.add_argument("User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, "
                     "like Gecko) Chrome/93.0.4577.82")
options.add_argument("lang=ko_KR")
driver = webdriver.Chrome(chd, options=options)

driver.get("https://anime.onnada.com/")

# 엑셀 파일 존재 여부 확인 후, 없으면 생성
xl = "ani_update.xlsx"
if os.path.isfile(xl):
    # 엑셀 파일 불러오기
    work_book = load_workbook(xl)
    sheet = work_book['Sheet']
else:
    # 엑셀 파일 생성
    work_book = Workbook()
    sheet = work_book.active
    sheet.column_dimensions["I"].width = 20
    sheet.column_dimensions["J"].width = 20
    sheet.column_dimensions["K"].width = 20
    sheet.column_dimensions["L"].width = 20

    sheet['I1'] = "1차 분기"
    sheet['I1'].font = Font(name="나눔고딕", color="000000", bold=True)
    sheet['I1'].alignment = Alignment(horizontal='center')
    sheet['J1'] = "1차 체크 링크"
    sheet['J1'].font = Font(name="나눔고딕", color="000000", bold=True)
    sheet['J1'].alignment = Alignment(horizontal='center')
    
    sheet['K1'] = "2차 체크 제목"
    sheet['K1'].font = Font(name="나눔고딕", color="000000", bold=True)
    sheet['K1'].alignment = Alignment(horizontal='center')
    sheet['L1'] = "2차 체크 링크"
    sheet['L1'].font = Font(name="나눔고딕", color="000000", bold=True)
    sheet['L1'].alignment = Alignment(horizontal='center')

    sheet.column_dimensions["A"].width = 55
    sheet.column_dimensions["B"].width = 55
    sheet.column_dimensions["C"].width = 55
    sheet.column_dimensions["D"].width = 30
    sheet.column_dimensions["E"].width = 30
    sheet.column_dimensions["H"].width = 30

    sheet['A1'] = "제목"
    sheet['A1'].font = Font(name="나눔고딕", color="000000", bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet['B1'] = "장르"
    sheet['B1'].font = Font(name="나눔고딕", color="000000", bold=True)
    sheet['B1'].alignment = Alignment(horizontal='center')
    sheet['C1'] = "태그"
    sheet['C1'].font = Font(name="나눔고딕", color="000000", bold=True)
    sheet['C1'].alignment = Alignment(horizontal='center')
    sheet['D1'] = "방영일"
    sheet['D1'].font = Font(name="나눔고딕", color="000000", bold=True)
    sheet['D1'].alignment = Alignment(horizontal='center')
    sheet['E1'] = "대표이미지 파일명"
    sheet['E1'].font = Font(name="나눔고딕", color="000000", bold=True)
    sheet['E1'].alignment = Alignment(horizontal='center')
    sheet['H1'] = "링크"
    sheet['H1'].font = Font(name="나눔고딕", color="000000", bold=True)
    sheet['H1'].alignment = Alignment(horizontal='center')

    work_book.save(xl)
    work_book.close()

    # 엑셀 파일 불러오기
    work_book = load_workbook(xl)
    sheet = work_book['Sheet']
print("\n★ 0) 엑셀 파일 불러오기 완료 ★")

# 업데이트 로그(Update Log)파일 불러오기
update_file = open("update_log.txt", 'r+')  # r : read, w : write, a : append
update_list = update_file.read().split("\n")
update_file.close()
print("\n★ 1) 업데이트 로그 파일 불러오기 완료 ★")

# 크롤링 할 내역이 있는지 확인
find_date_xpath = "/html/body/div/div/div/div/ul/li/a"
find_date = driver.find_elements_by_xpath(find_date_xpath)

# 1차[I, J] : 애니메이션 분기 크롤링
temp_update_check_list = []
update_crawl_dict = {}
for index, da in enumerate(find_date):
    update_crawl_dict[da.text] = str(da.get_attribute("href"))
    temp_update_check_list.append(da.text)

# 업데이트 필요성 체크
pass_count = 0
update_add_list = []
for vv in temp_update_check_list:
    if vv not in update_list:
        pass_count = 1
        update_add_list.append(vv)
if pass_count == 0:
    print("\n★ 이미 최신 업데이트입니다! ★")
    quit()

# 엑셀파일에 업데이트 분기 정보 넣기
for no, ddd in enumerate(update_add_list):
    sheet["I"+str(no + 2)].value = ddd  # 분기
    sheet["J"+str(no + 2)].value = update_crawl_dict[ddd]  # 링크
    print("\n 업데이트할 분기 : ", ddd)

work_book.save(xl)
print("\n★ 2) 애니메이션 분기 크롤링 완료 ★")
print("\n\n - 2초 대기 - 엑셀파일 저장 후 재로딩\n\n")
time.sleep(2)

# 2차[K, L] : 애니메이션 목록 크롤링
print("★ ... 애니메이션 목록 크롤링 중... ★")
work_book = load_workbook(xl)
sheet = work_book['Sheet']
genre_count = 1
detail_link_xpath = "/html/body/div/div/div/div/div/ul/div/li/p/a"
count = 2
for no in range(2, len(sheet["I"]) + 1):
    driver.get(sheet["J"+str(no)].value)
    detail_link = driver.find_elements_by_xpath(detail_link_xpath)
    for kk in detail_link:
        if kk.text != "":
            sheet["K"+str(count)].value = kk.text  # 제목
            sheet["L"+str(count)].value = kk.get_attribute("href")  # 링크
            count += 1

work_book.save(xl)
print("\n★ 3) 애니메이션 목록 크롤링 완료 ★")
print("\n\n - 2초 대기 - 엑셀파일 저장 후 재로딩\n\n")
time.sleep(2)

# 애니메이션 1차 크롤링
print("★ ... 애니메이션 1차 크롤링 중... ★")
work_book = load_workbook(xl)
sheet = work_book['Sheet']
title_xpath = "/html/body/div/div/div/article/div/h1"
genre_check_xpath = "/html/body/div/div/div/article/div/div/p/span[1]"
image_xpath = "/html/body/div/div/div/article/div/div/div/a/img"
content_xpath = "//*[@id='animeContents']"

for no in range(2, len(sheet["K"]) + 1):
    try:
        # 크롤링 주소 가져오기
        driver.get(sheet["L"+str(no)].value)
        time.sleep(3)
        # 링크 시트 삽입
        sheet["H"+str(no)].value = sheet["L"+str(no)].value

        # XPATH 찾기
        title_tag = driver.find_elements_by_xpath(title_xpath)
        genre_check_tag = driver.find_elements_by_xpath(genre_check_xpath)
        image_tag = driver.find_elements_by_xpath(image_xpath)

        # 제목 크롤링
        for title in title_tag:
            sheet["A"+str(no)].value = title.text
            print("\n [ ", str(no), " ]", title.text)

        # 장르, 방영일 크롤링
        genre_count = 1
        date_count = 1
        for aa in genre_check_tag:
            if aa.text == "장르":
                genre_xpath = "/html/body/div/div/div/article/div/div/p[" + str(genre_count) + "]/span[2]"
                genre_tag = driver.find_elements_by_xpath(genre_xpath)
                for genre in genre_tag:
                    sheet["B" + str(no)].value = genre.text

            if aa.text == "방영일":
                date_xpath = "/html/body/div/div/div/article/div/div/p[" + str(date_count) + "]/span[2]"
                date_tag = driver.find_elements_by_xpath(date_xpath)
                for date in date_tag:
                    sheet["D" + str(no)].value = date.text

            genre_count += 1
            date_count += 1

        # 줄거리 크롤링
        content_tag = driver.find_elements_by_xpath(content_xpath)
        if len(content_tag) > 0:
            for kk in content_tag:
                if str(kk.text).replace(" ", "") == "-":  # 줄거리 X
                    sheet["F" + str(no)].value = ""
                else:  # 줄거리 O
                    sheet["F" + str(no)].value = kk.text
        else:  # 줄거리 X
            sheet["F" + str(no)].value = ""

        # 대표 이미지 크롤링 및 다운로드
        for image in image_tag:
            # 시간으로 파일명 지정
            now_time = datetime.datetime.now()
            front_time = str(now_time.year) + str(now_time.month) + str(now_time.day)
            back_time = str(now_time.hour) + str(now_time.minute) + str(now_time.second) + str(now_time.microsecond)
            image_file_name = front_time + back_time + ".jpg"

            # 다운받을 이미지 링크
            url = image.get_attribute("src")

            # 이미지 다운로드
            urllib.request.urlretrieve(url, "../static/image/" + image_file_name)
            time.sleep(5)

            sheet["E" + str(no)].value = image_file_name

    except StaleElementReferenceException:
        print("StaleElementReferenceException")
    except InvalidArgumentException:
        print("InvalidArgumentException")
    except NoSuchElementException:
        print("NoSuchElementException")
    except TimeoutException:
        print("TimeoutException")
        driver.quit()

        time.sleep(5)
        options = webdriver.ChromeOptions()
        options.add_argument("headless")
        options.add_argument("window-size=1920x1080")
        options.add_argument("disable-gpu")
        options.add_argument("User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, "
                             "like Gecko) Chrome/93.0.4577.82")
        options.add_argument("lang=ko_KR")
        driver = webdriver.Chrome(chd, options=options)

    except WebDriverException:
        print("WebDriverException")
    except IndexError:
        print("IndexError")
    except:
        print("Error 발생!")

work_book.save(xl)
print("\n★ 4) 애니메이션 1차 크롤링 완료 ★")
print("\n\n - 2초 대기 - 엑셀파일 저장 후 재로딩\n\n")
time.sleep(2)

# 애니메이션 2차 크롤링
print("★ ... 애니메이션 2차 크롤링 중... ★")
driver.get("https://www.chuing.net/db/search.php?cdbsearch=asdfasdf")
search_xpath = "//*[@id='SjestForm']/div/div[1]/input"
result_a_xpath = "/html/body/div[5]/div[2]/div/div[2]/div/div[4]/div[1]"
work_book = load_workbook(xl)
sheet = work_book['Sheet']

for no in range(2, len(sheet["A"]) + 1):
    try:
        search_tag = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, search_xpath)))
        search_tag.clear()

        search_tag.send_keys(sheet["A" + str(no)].value)
        search_tag.send_keys(Keys.ENTER)

        result_a = driver.find_elements_by_xpath(result_a_xpath)
        result_total = []
        sheet_b = sheet["B" + str(no)].value

        if len(result_a) > 0:
            for a in result_a:
                for kkk in str(a.text).split("장르: ")[1].split(","):
                    result_total.append(kkk)
        if len(result_total) > 0:
            if sheet_b is not None and sheet_b != "" and sheet_b != "None":
                sheet["B" + str(no)].value = sheet_b + ", " + ", ".join(result_total)
            else:
                sheet["B" + str(no)].value = ", ".join(result_total)

        print("[ ", str(no), " ] ", sheet["A" + str(no)].value, "[ ", sheet["B" + str(no)].value, " ]")
        time.sleep(3)

    except StaleElementReferenceException:
        print("StaleElementReferenceException")
    except InvalidArgumentException:
        print("InvalidArgumentException")
    except TimeoutException:
        print("TimeoutException")
    except NoSuchElementException:
        print("NoSuchElementException")
    except WebDriverException:
        print("WebDriverException")
    except IndexError:
        print("IndexError")
driver.quit()

# 업데이트 엑셀 파일 생성
work_book.save(xl)
print("\n★ 6) 업데이트 엑셀 파일 생성 완료 ★")

# 업데이트 로그(Update Log) 작성
success_update_file = open("update_log.txt", 'a')  # r : read, w : write, a : append
if len(update_add_list) > 0:
    for mm in update_add_list:
        success_update_file.write("\n" + mm)
success_update_file.close()
print("\n★ 7) 업데이트 로그 파일 작성 완료 ★")

print("\n★ ! Success !")
