from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time

# https://anime.onnada.com/

chd = 'C:/dev_files/chd/chd.exe'
options = webdriver.ChromeOptions()
options.add_argument("headless")
options.add_argument("window-size=1920x1080")
options.add_argument("disable-gpu")
options.add_argument("User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, "
                     "like Gecko) Chrome/93.0.4577.82")
options.add_argument("lang=ko_KR")
driver = webdriver.Chrome(chd, options=options)

# 엑셀파일 생성
save_book = Workbook()
# 시트 입력
save_sheet = save_book.active
save_sheet.column_dimensions["A"].width = 55
save_sheet.column_dimensions["B"].width = 55

save_sheet['A1'] = "제목"
save_sheet['A1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['A1'].alignment = Alignment(horizontal='center')
save_sheet['B1'] = "링크"
save_sheet['B1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['B1'].alignment = Alignment(horizontal='center')

# 리스트 불러오기
load_book = load_workbook("anime_find_date.xlsx")
load_sheet = load_book['Sheet']

genre_count = 1
detail_link_xpath = "/html/body/div/div/div/div/div/ul/div/li/p/a"

count = 2
for no in range(2, len(load_sheet["A"])):
    driver.get(load_sheet["A"+str(no)].value)
    detail_link = driver.find_elements_by_xpath(detail_link_xpath)
    for kk in detail_link:
        if kk.text != "":
            save_sheet["A"+str(count)].value = kk.text
            save_sheet["B"+str(count)].value = kk.get_attribute("href")
            count += 1
            print("[ ", str(count), " ] 제목 : ", kk.text)

driver.quit()

save_book.save("anime_detail_link.xlsx")

