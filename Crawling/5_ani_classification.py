from openpyxl import Workbook, load_workbook

print("\n 0 : 문제 확인")
print(" 1 : 문자열 치환")
print(" 2 : 중복 제거")
print(" 3 : 장르(Genre)분류 후, 태그(Tag) 이동")
select = input("\n 선택 : ")
xl = "ani_detail.xlsx"

# 문제 확인
if select == "0":
    work_book = load_workbook(xl)
    sheet = work_book['Sheet']
    check_list = []
    etc_list = []
    none_list = []
    blank_list = []
    image_blank_list = []
    temp_genre_list = []
    for no in range(2, len(sheet["A"]) + 1):
        # 문제 확인 : 기타
        if sheet["B" + str(no)].value is not None:
            for kk in sheet["B" + str(no)].value.split(", "):
                if kk == "기타":
                    temp_str = "[ " + str(no) + " ] " + sheet["A" + str(no)].value
                    etc_list.append(temp_str)
                check_list.append(kk)

        # 문제 확인 : None
        if sheet["B"+str(no)].value == "None" or sheet["B"+str(no)].value is None or sheet["B"+str(no)].value == "":
            temp_none = "None : [ " + str(no) + " ] " + sheet["A"+str(no)].value
            none_list.append(temp_none)

        # 문제 확인 : 공백
        if sheet["B"+str(no)].value == "" or sheet["B"+str(no)].value == " " or sheet["B"+str(no)].value is None:
            temp_blank = "공백(Blank) : [ " + str(no) + " ] " + sheet["A" + str(no)].value
            blank_list.append(temp_blank)

        # 문제 확인 : 장르가 1개인데, 대표 장르 목록에 포함되어 있지 않은 경우
        genre_list = ["스포츠", "추리", "코미디", "성전환", "먼치킨", "모험", "SF", "하렘", "드라마", "이세계",
                      "일상", "학원", "로맨스", "미스테리", "치유", "판타지", "메카", "스릴러", "19금",
                      "공포", "공상과학", "초자연", "음악", "아동", "액션", "BL", "백합", "시대"]
        if sheet["B" + str(no)].value is not None:
            for kk in sheet["B" + str(no)].value.split(", "):
                if len(sheet["B" + str(no)].value.split(", ")) == 1 and kk not in genre_list:
                    temp_str = "[ " + str(no) + " ] " + sheet["A" + str(no)].value + "[ " + \
                               sheet["B" + str(no)].value + " ]"
                    temp_genre_list.append(temp_str)

    check_list = list(set(check_list))
    print(" 장르 목록 : ", check_list)
    if len(etc_list) > 0:
        print("\n\n 기타 목록 ")
        for etc in etc_list:
            print(etc)
    if len(none_list) > 0:
        print("\n\n None 목록 ")
        for none in none_list:
            print(none)
    if len(blank_list) > 0:
        print("\n\n 공백(Blank) 목록 ")
        for blank in blank_list:
            print(blank)
    if len(temp_genre_list) > 0:
        print("\n\n 장르가 아닌 목록 ")
        for genre in temp_genre_list:
            print(genre)

# 문자열 치환
if select == "1":
    work_book = load_workbook(xl)
    sheet = work_book['Sheet']
    for no in range(2, len(sheet["A"]) + 1):
        sheet["B" + str(no)].value = str(sheet["B" + str(no)].value)\
            .replace("롤플레잉", "delete").replace("아케이드", "delete").replace("퍼즐/카드", "delete")\
            .replace("치유물", "치유")\
            .replace("에찌", "19금").replace("헨타이", "19금")\
            .replace("소녀Ai", "소녀")\
            .replace("어드벤처", "모험")\
            .replace("아이들", "아동").replace("아동물", "아동")\
            .replace("TS", "성전환")\
            .replace("소년사랑", "소년, 로맨스")\
            .replace("학교", "학원")\
            .replace("메카닉", "메카")

        sheet["B" + str(no)].value = str(sheet["B" + str(no)].value)\
            .replace(", delete", "").replace("delete, ", "")
    work_book.save(xl)

# 중복 제거
if select == "2":
    work_book = load_workbook(xl)
    sheet = work_book['Sheet']
    for no in range(2, len(sheet["A"]) + 1):
        temp_sheet = list(set(str(sheet["B" + str(no)].value).split(", ")))
        sheet["B" + str(no)].value = ", ".join(temp_sheet)

    work_book.save(xl)

    for no in range(2, len(sheet["A"]) + 1):
        if sheet["B"+str(no)].value == "None" or sheet["B"+str(no)].value is None or sheet["B"+str(no)].value == "":
            print("None : [ ", str(no), " ] ", sheet["A"+str(no)].value)

# 장르(Genre)를 제외한 나머지 태그(Tag)로 이동
if select == "3":
    genre_list = ["스포츠", "추리", "코미디", "성전환", "먼치킨", "모험", "SF", "하렘", "드라마", "이세계",
                  "일상", "학원", "로맨스", "미스테리", "치유", "판타지", "메카", "스릴러", "19금",
                  "공포", "공상과학", "초자연", "음악", "아동", "액션", "BL", "백합", "시대"]
    work_book = load_workbook(xl)
    sheet = work_book['Sheet']
    for no in range(2, len(sheet["A"]) + 1):
        tag_data = []
        genre_data = []
        for kk in sheet["B" + str(no)].value.split(", "):
            if kk not in genre_list:
                tag_data.append(kk)
            else:
                genre_data.append(kk)

        sheet["B" + str(no)].value = ", ".join(genre_data)
        sheet["C" + str(no)].value = ", ".join(tag_data)

    work_book.save(xl)
