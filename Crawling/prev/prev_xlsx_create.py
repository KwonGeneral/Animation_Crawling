
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

save_book = Workbook()

# 시트 입력
save_sheet = save_book.active
save_sheet.column_dimensions["A"].width = 55
save_sheet.column_dimensions["B"].width = 55
save_sheet.column_dimensions["C"].width = 55
save_sheet.column_dimensions["D"].width = 30
save_sheet.column_dimensions["E"].width = 30
save_sheet.column_dimensions["H"].width = 30

save_sheet['A1'] = "제목"
save_sheet['A1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['A1'].alignment = Alignment(horizontal='center')
save_sheet['B1'] = "장르"
save_sheet['B1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['B1'].alignment = Alignment(horizontal='center')
save_sheet['C1'] = "태그"
save_sheet['C1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['C1'].alignment = Alignment(horizontal='center')
save_sheet['D1'] = "제작년도"
save_sheet['D1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['D1'].alignment = Alignment(horizontal='center')
save_sheet['E1'] = "대표이미지"
save_sheet['E1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['E1'].alignment = Alignment(horizontal='center')
save_sheet['H1'] = "링크"
save_sheet['H1'].font = Font(name="나눔고딕", color="000000", bold=True)
save_sheet['H1'].alignment = Alignment(horizontal='center')

xlsx_list = ["ani_a.xlsx", "ani_b.xlsx", "ani_c.xlsx", "ani_d.xlsx", "ani_e.xlsx", "ani_f.xlsx", "ani_g.xlsx",
             "ani_h.xlsx", "ani_i.xlsx", "ani_j.xlsx", "ani_k.xlsx", "ani_l.xlsx", "ani_m.xlsx", "ani_n.xlsx",
             "ani_o.xlsx", "ani_p.xlsx"]

count = 2
for xl in xlsx_list:
    load_book = load_workbook(xl)
    load_sheet = load_book['Sheet']
    for no in range(2, len(load_sheet["A"])):
        save_sheet["A" + str(count)].value = load_sheet["A" + str(no)].value
        save_sheet["A" + str(count)].font = Font(name="나눔고딕", color="000000")

        save_sheet["B" + str(count)].value = load_sheet["B" + str(no)].value
        save_sheet["B" + str(count)].font = Font(name="나눔고딕", color="000000")

        save_sheet["C" + str(count)].value = load_sheet["C" + str(no)].value
        save_sheet["C" + str(count)].font = Font(name="나눔고딕", color="000000")

        save_sheet["H" + str(count)].value = load_sheet["H" + str(no)].value
        save_sheet["H" + str(count)].font = Font(name="나눔고딕", color="50BCDF")
        count += 1


save_book.save("prev_total.xlsx")

