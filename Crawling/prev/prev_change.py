from openpyxl import load_workbook

# xlsx_list = ["ani_a.xlsx", "ani_b.xlsx", "ani_c.xlsx", "ani_d.xlsx", "ani_e.xlsx", "ani_f.xlsx", "ani_g.xlsx",
#              "ani_h.xlsx", "ani_i.xlsx", "ani_j.xlsx", "ani_k.xlsx", "ani_l.xlsx", "ani_m.xlsx", "ani_n.xlsx",
#              "ani_o.xlsx", "ani_p.xlsx"]

# xl = xlsx_list[0]
# work_book = load_workbook(xl)
# sheet = work_book['Sheet']

# # 중복 제거
# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"]) + 1):
#         temp_sheet = list(set(str(sheet["B" + str(no)].value).split(", ")))
#         sheet["B" + str(no)].value = ", ".join(temp_sheet)
#
#         work_book.save(xl)

# None 제거
# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"]) + 1):
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value).replace("None", "")
#
#         work_book.save(xl)

# # 치환 1)
# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"]) + 1):
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value)\
#             .replace("애니메이션", "delete").replace("어드벤처", "모험").replace("물", "")\
#             .replace("희극", "코미디").replace("거대로봇", "메카").replace("만화", "delete")\
#             .replace("롤플레잉", "delete").replace("게임", "delete").replace("Sports manga", "스포츠")\
#             .replace("액션/어드벤처", "액션, 모험").replace("액션/모험", "액션, 모험").replace("Girls with guns", "delete")\
#             .replace("소설", "delete").replace("None", "").replace("에찌", "19금")\
#             .replace("전기", "delete").replace("노블", "delete").replace("헨타이", "19금")\
#             .replace("그라제니", "delete").replace("초자연적 현상", "delete").replace("성년", "19금")\
#             .replace("슈퍼파워", "액션").replace("소년", "delete").replace("월드시네마", "delete")\
#             .replace("영화", "delete").replace("소녀Ai", "delete").replace("비극", "슬픔")\
#             .replace("소년사랑", "로맨스").replace("테크느와르", "느와르").replace("어린이", "delete")\
#             .replace("소녀", "delete")\
#             .replace("RPG", "delete")\
#             .replace("청년", "delete")\
#             .replace("슈퍼히어로", "히어로")\
#             .replace("장르", "delete")\
#             .replace("방송", "delete")\
#             .replace("일본", "delete")\
#             .replace("역사허구", "delete")\
#             .replace("텔레비전", "delete")\
#             .replace("단편", "delete")\
#             .replace("시뮬레이션", "delete")\
#             .replace("에로게", "delete")\
#             .replace("무비", "delete")\
#             .replace("학교", "학원")\
#             .replace("아니메", "delete")\
#             .replace("VR", "delete")\
#             .replace("연애 delete", "로맨스")\
#             .replace("풍자", "패러디")\
#             .replace("순정", "로맨스")\
#             .replace("로봇", "delete")\
#             .replace("섹스 코미디", "19금")\
#             .replace("과학 판타지", "판타지")\
#             .replace("아이들", "delete")\
#             .replace("노벨", "delete")\
#             .replace("Sex and nudity in video games", "delete")\
#             .replace("서스펜스", "스릴러")\
#             .replace("성장 스토리", "왕도")\
#             .replace("여성", "delete")\
#             .replace("고딕", "delete")\
#             .replace("전차 코미디", "delete")\
#             .replace("LGBT", "delete")\
#             .replace("에로티카", "delete")\
#             .replace("피카레스크", "delete")\
#             .replace("리얼로봇", "메카")\
#             .replace("미니시리즈", "delete")\
#             .replace("4컷", "delete")
#
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value)\
#             .replace("일상의 delete", "일상").replace("일상 delete", "일상")\
#             .replace("판타지 delete", "판타지").replace("판타지의 delete", "판타지")\
#             .replace("메카의 delete", "메카").replace("메카 delete", "메카")\
#             .replace("로봇 delete", "메카").replace("로봇의 delete", "메카")\
#             .replace("코미디 delete", "코미디").replace("코미디의 delete", "코미디")\
#             .replace("우주의 delete", "우주").replace("우주의 delete", "우주")\
#             .replace("미스터리 delete", "미스터리").replace("미스터리의 delete", "미스터리")\
#             .replace("모험의 delete", "모험").replace("모험 delete", "모험")\
#             .replace("동화의 delete", "동화").replace("동화 delete", "동화")\
#             .replace("공포 delete", "공포").replace("공포의 delete", "공포").replace("공포 delete의 delete", "공포")\
#             .replace("SF의 delete", "SF").replace("SF delete", "SF").replace("SF delete의 delete", "SF")\
#             .replace("심리의 delete", "심리").replace("심리 delete", "심리")\
#             .replace("마법delete", "마법소녀")\
#             .replace("군사의 delete", "군사").replace("군사 delete", "군사")\
#             .replace("드라마의 delete", "드라마").replace("드라마 delete", "드라마")\
#             .replace("액션 delete", "액션").replace("액션의 delete", "액션")\
#             .replace("로맨스 delete", "로맨스").replace("로맨스의 delete", "로맨스")\
#             .replace("개그 delete", "개그").replace("개그의 delete", "개그")\
#             .replace("성장", "왕도")\
#             .replace("학원 delete", "학원").replace("학원의 delete", "학원")\
#             .replace("백합 delete", "백합").replace("백합의 delete", "백합")\
#             .replace("스포츠 delete", "스포츠").replace("스포츠의 delete", "스포츠")\
#             .replace("다크판타지 delete", "다크판타지").replace("다크판타지의 delete", "다크판타지")\
#             .replace("다크 판타지 delete", "다크 판타지").replace("다크 판타지의 delete", "다크 판타지")\
#             .replace("공상과학 delete", "공상과학").replace("공상과학의 delete", "공상과학")\
#             .replace("순정 delete", "로맨스").replace("순정의 delete", "로맨스")\
#             .replace("초자연 delete", "초자연").replace("초자연의 delete", "초자연")\
#             .replace("요리 delete", "요리").replace("요리의 delete", "요리")\
#             .replace("음악 delete", "음악").replace("음악의 delete", "음악")\
#             .replace("호러 delete", "호러").replace("호러의 delete", "호러")\
#             .replace("일상의delete", "일상").replace("일상delete", "일상") \
#             .replace("판타지delete", "판타지").replace("판타지의delete", "판타지") \
#             .replace("메카의delete", "메카").replace("메카delete", "메카") \
#             .replace("로봇delete", "메카").replace("로봇의delete", "메카") \
#             .replace("코미디delete", "코미디").replace("코미디의delete", "코미디") \
#             .replace("우주의delete", "우주").replace("우주의delete", "우주") \
#             .replace("미스터리delete", "미스터리").replace("미스터리의delete", "미스터리") \
#             .replace("모험의delete", "모험").replace("모험delete", "모험") \
#             .replace("동화의delete", "동화").replace("동화delete", "동화") \
#             .replace("공포delete", "공포").replace("공포의delete", "공포")\
#             .replace("SF의delete", "SF").replace("SFdelete", "SF") \
#             .replace("심리의delete", "심리").replace("심리delete", "심리") \
#             .replace("군사의delete", "군사").replace("군사delete", "군사") \
#             .replace("드라마의delete", "드라마").replace("드라마delete", "드라마") \
#             .replace("액션delete", "액션").replace("액션의delete", "액션") \
#             .replace("로맨스delete", "로맨스").replace("로맨스의delete", "로맨스") \
#             .replace("개그delete", "개그").replace("개그의delete", "개그") \
#             .replace("학원delete", "학원").replace("학원의delete", "학원") \
#             .replace("백합delete", "백합").replace("백합의delete", "백합") \
#             .replace("스포츠delete", "스포츠").replace("스포츠의delete", "스포츠") \
#             .replace("다크판타지delete", "다크판타지").replace("다크판타지의delete", "다크판타지") \
#             .replace("다크 판타지delete", "다크 판타지").replace("다크 판타지의delete", "다크 판타지") \
#             .replace("공상과학delete", "공상과학").replace("공상과학의delete", "공상과학") \
#             .replace("순정delete", "로맨스").replace("순정의delete", "로맨스") \
#             .replace("초자연delete", "초자연").replace("초자연의delete", "초자연") \
#             .replace("요리delete", "요리").replace("요리의delete", "요리") \
#             .replace("음악delete", "음악").replace("음악의delete", "음악") \
#             .replace("호러delete", "호러").replace("호러의delete", "호러")
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value).replace(", delete", "").replace("delete ,", "")
#
#         work_book.save(xl)

# 치환 2)
# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"]) + 1):
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value)\
#             .replace("로봇 delete", "메카").replace("로봇의 delete", "메카")\
#             .replace("로봇delete", "메카").replace("로봇의delete", "메카")\
#             .replace("메카 delete", "메카").replace("메카의 delete", "메카")\
#             .replace("메카delete", "메카").replace("메카의delete", "메카")\
#             .replace("모험 delete", "모험").replace("모험의 delete", "모험")\
#             .replace("모험delete", "모험").replace("모험의delete", "모험")\
#             .replace("밀리터리 delete", "밀리터리").replace("밀리터리의 delete", "밀리터리")\
#             .replace("밀리터리delete", "밀리터리").replace("밀리터리의delete", "밀리터리")\
#             .replace("배틀로얄 delete", "배틀로얄").replace("배틀로얄의 delete", "배틀로얄")\
#             .replace("배틀로얄delete", "배틀로얄").replace("배틀로얄의delete", "배틀로얄")\
#             .replace("스팀펑크 delete", "스팀펑크").replace("스팀펑크의 delete", "스팀펑크")\
#             .replace("스팀펑크delete", "스팀펑크").replace("스팀펑크의delete", "스팀펑크")\
#             .replace("리얼 delete", "delete").replace("리얼의 delete", "delete")\
#             .replace("리얼delete", "delete").replace("리얼의delete", "delete")\
#             .replace("범죄delete delete", "범죄")\
#             .replace("범죄 delete", "범죄").replace("범죄의 delete", "범죄")\
#             .replace("범죄delete", "범죄").replace("범죄의delete", "범죄")\
#             .replace("택티컬 delete", "delete")\
#             .replace("액션 delete", "액션").replace("액션의 delete", "액션")\
#             .replace("액션delete", "액션").replace("액션의delete", "액션")\
#             .replace("컴퓨터 delete delete", "delete")\
#             .replace("학원 delete", "학원").replace("학원의 delete", "학원")\
#             .replace("학원delete", "학원").replace("학원의delete", "학원")\
#             .replace("비주얼 delete", "delete")
#
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value).replace(", delete", "").replace("delete, ", "")
#
#         work_book.save(xl)


# 치환 3)
# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"]) + 1):
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value)\
#             .replace("어드벤쳐", "모험")\
#             .replace("사이언스", "delete")\
#             .replace("심리스릴러", "delete")\
#             .replace("도시 판타지", "판타지")\
#             .replace("어반 판타지", "판타지")\
#             .replace("구루메", "음식")\
#             .replace("세카이계", "세카이")\
#             .replace("멜로", "delete")\
#             .replace("그래픽", "delete")\
#             .replace("이세계 전생", "이세계")\
#             .replace("픽션", "delete")\
#             .replace("퀴즈", "delete")\
#             .replace("카드", "delete")\
#             .replace("멜로", "delete")\
#             .replace("서브컬처", "delete")\
#             .replace("수명", "delete")\
#             .replace("아동 문학", "delete")\
#             .replace("도덕극", "delete")\
#             .replace("사회 평론", "delete")\
#             .replace("비화", "delete")\
#             .replace("아케이드", "delete")\
#             .replace("이세계의 판타지", "이세계")\
#             .replace("웹코믹", "delete").replace("뮤지컬", "delete").replace("24분", "delete")
#
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value).replace(", delete", "").replace("delete, ", "")
#
#         work_book.save(xl)

# 치환 4)
# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"]) + 1):
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value)\
#             .replace("능력자 배틀", "이능력, 배틀")\
#             .replace("판타지배틀", "판타지, 배틀")\
#             .replace("에픽 판타지", "판타지")\
#             .replace("코믹 판타지", "판타지")\
#             .replace("음양의 판타지", "판타지")\
#             .replace("역사 판타지", "판타지")\
#             .replace("신화적 판타지", "판타지")\
#             .replace("오버홀 판타지", "판타지")\
#             .replace("로드 판타지", "판타지")\
#             .replace("판타지T모험", "판타지, 모험")\
#             .replace("코미디 드라마", "코미디, 드라마")\
#             .replace("러브 코미디", "로맨틱 코미디")\
#             .replace("로맨스 코미디", "로맨틱 코미디")\
#             .replace("신코미디", "코미디")\
#             .replace("스포츠 코미디", "스포츠, 코미디")\
#             .replace("호러 코미디", "호러, 코미디")\
#             .replace("마법소녀의 코미디", "마법소녀, 코미디")\
#             .replace("음악 코미디", "음악, 코미디")\
#             .replace("부조리 코미디", "코미디")\
#             .replace("명랑코미디", "코미디")\
#             .replace("리듬 코미디", "코미디")\
#             .replace("이세계의 코미디", "이세계, 코미디")\
#             .replace("코미디사랑", "코미디, 로맨스")\
#             .replace("포스트 아포칼립스", "아포칼립스")\
#             .replace("진행형 격투 액션", "격투, 액션")\
#             .replace("검과 마법", "마법")\
#             .replace("수퍼내추럴 드라마", "드라마")\
#             .replace("이세계 전이", "이세계")\
#             .replace("축구의 스포츠", "축구, 스포츠")\
#             .replace("포르노", "19금")\
#             .replace("의학 드라마", "의학, 드라마")\
#             .replace("공도 레이싱", "레이싱")\
#             .replace("사이버펑크 파생", "사이버펑크")\
#             .replace("밀리터리 드라마", "밀리터리, 드라마")\
#             .replace("밀리터리 SF", "밀리터리, SF")\
#             .replace("성전환", "TS")\
#             .replace("비현실적 유머", "개그")\
#             .replace("장편 범죄", "범죄")\
#             .replace("철학", "delete")\
#             .replace("포니캐년", "delete")\
#             .replace("정치", "delete")\
#             .replace("디젤펑크의 드라마", "드라마")\
#             .replace("환상문학", "delete")\
#             .replace("현대의 치유", "치유")\
#             .replace("로맨스/로맨스", "로맨스")\
#             .replace("초현실주의 드라마", "드라마")\
#             .replace("사랑", "delete")\
#             .replace("슈퍼드라마", "드라마")\
#             .replace("춤", "댄스")\
#             .replace("SF의 테크누아르", "SF, 느와르")\
#             .replace("마법가", "마법")\
#             .replace("19금의 학원", "19금, 학원")\
#             .replace("가족의 드라마", "가족, 드라마")\
#             .replace("문학", "delete")\
#             .replace("초현실주의", "delete")\
#             .replace("라이트", "delete")\
#             .replace("심리호러", "심리, 호러")\
#             .replace("코믹", "코미디")\
#             .replace("빅토리아 시대", "delete")\
#             .replace("악마 호러", "악마, 호러")\
#             .replace("배틀의 범죄", "배틀, 범죄")\
#             .replace("퍼즐/하렘", "하렘")\
#             .replace("하드보일드", "delete")\
#             .replace("POP", "delete")\
#             .replace("록", "delete")\
#             .replace("서바이벌 호러", "호러")\
#             .replace("인디 호러", "호러")\
#             .replace("유령 이야기", "호러")\
#             .replace("히어로SF", "히어로, SF")\
#             .replace("뱀파이어 공포 드라마", "뱀파이어, 공포, 드라마")\
#             .replace("러브 코미디 괴도 액션", "로맨틱 코미디, 괴도, 액션")\
#             .replace("히어로액션", "히어로, 액션")\
#             .replace("우주서부드라마", "우주, 드라마")\
#             .replace("초자연적 미스터리", "초자연, 미스터리")\
#             .replace("네오누아르", "느와르")\
#             .replace("테크누아르", "느와르")\
#             .replace("누와르", "느와르")\
#             .replace("프로그램", "delete")\
#             .replace("디멘시아", "delete")
#
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value).replace(", delete", "").replace("delete, ", "")
#
#         work_book.save(xl)


# 치환 5)
# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"]) + 1):
#         sheet["B" + str(no)].value = str(sheet["B" + str(no)].value)\
#             .replace("판타지의 드라마", "판타지, 드라마")\
#             .replace("19금의 로맨틱 코미디", "19금, 로맨틱, 코미디")\
#             .replace("에로틱 코미디", "19금, 코미디")\
#             .replace("연애 모험", "로맨스, 모험")\
#             .replace("마법소녀의 마법", "마법소녀, 마법")\
#             .replace("러브코미디", "로맨틱 코미디")\
#             .replace("판타지T마법", "판타지, 마법")\
#             .replace("드라마 로맨스", "드라마, 로맨스")\
#             .replace("우주서부코미디", "우주, 코미디")\
#             .replace("판타지 액션", "판타지, 액션")\
#             .replace("SF 코미디", "SF, 코미디")\
#             .replace("드라마 모험", "드라마, 모험")\
#             .replace("요괴 호러", "요괴, 호러")\
#             .replace("코미디 밀리터리", "코미디, 밀리터리")\
#             .replace("9금", "19금")\
#             .replace("19금미스터리", "19금, 미스터리")\
#             .replace("드라마의 로맨스", "드라마, 로맨스")\
#             .replace("코메디 등장인", "delete")\
#             .replace("오토메이트의 판타지", "판타지")\
#             .replace("스파이 미스터리", "스파이, 미스터리")\
#             .replace("SF의 모험", "SF, 모험")\
#             .replace("쿠킹 판타지", "음식, 판타지")\
#             .replace("심리의 미스터리", "심리, 미스터리")\
#             .replace("모험 메카", "모험, 메카")\
#             .replace("스릴러액션", "스릴러, 액션")\
#             .replace("학원 드라마", "학원, 드라마")\
#             .replace("로맨스일상", "로맨스, 일상")\
#             .replace("마녀의 로맨스", "로맨스")\
#             .replace("액션 모험", "액션, 모험")\
#             .replace("장편 드라마", "드라마")\
#             .replace("역사 마법", "역사, 마법")\
#             .replace("아포칼립스 SF", "아포칼립스, SF")\
#             .replace("배틀의 코미디", "배틀, 코미디")\
#             .replace("OST", "delete")\
#             .replace("노가미 타케시", "delete")\
#             .replace("전쟁의 드라마", "전쟁, 드라마")\
#             .replace("패밀리 드라마", "가족, 드라마")\
#             .replace("스포츠 왕도", "스포츠, 왕도")\
#             .replace("마법 모험", "마법, 모험")\
#             .replace("일상드라마", "일상, 드라마")\
#             .replace("학원 배틀", "학원, 배틀")\
#             .replace("SF의 공상과학", "SF, 공상과학")\
#             .replace("역사의 드라마", "역사, 드라마")\
#             .replace("스릴러 스릴러", "스릴러")\
#             .replace("코미디의 판타지", "코미디, 판타지")\
#             .replace("액션 월간 코미디 제논", "액션, 코미디")\
#             .replace("역사야오이", "역사")\
#             .replace("음악 비디오 일상", "음악, 일상")\
#             .replace("액션 스릴러", "액션, 스릴러")\
#             .replace("이능력 배틀", "이능력, 배틀")\
#             .replace("삼양출판사", "delete")\
#             .replace("야오이", "delete")\
#             .replace("모험모험", "모험, 모험")\
#             .replace("각본: 엔도 아키노리; 코나", "delete")\
#             .replace("카 치아키", "delete")\
#             .replace("개그의 밀리터리", "개그, 밀리터리")\
#             .replace("혼고 미츠루", "delete")\
#             .replace("법률 드라마", "드라마")\
#             .replace("마술적 사실주의", "delete")\
#             .replace("모험 코미디", "모험, 코미디")\
#             .replace("일상 치유", "일상, 치유")\
#             .replace("코미디 공상과학", "코미디, 공상과학")\
#             .replace("핵 앤드 슬래시", "delete")\
#             .replace("코미디 메카", "코미디, 메카")\
#             .replace("19금의 로맨스", "19금, 로맨스")\
#             .replace("로맨스 판타지", "로맨스, 판타지")\
#             .replace("드라마의 판타지", "드라마, 판타지")\
#             .replace("판타지 밀리터리", "판타지, 밀리터리")\
#             .replace("드라마 메카", "드라마, 메카")\
#             .replace("코미디의 모험", "코미디, 모험")\
#             .replace("SF 스릴러의 드라마", "SF, 스릴러, 드라마")\
#             .replace("호색로맨스 공상과학", "로맨스, 공상과학")\
#             .replace("역사 드라마", "역사, 드라마")\
#             .replace("코미디 로맨스", "코미디, 로맨스")\
#             .replace("드라마의 모험", "드라마, 모험")\
#             .replace("공포 호러", "공포, 호러")\
#             .replace("일상 코미디", "일상, 코미디")\
#             .replace("전쟁 이야기", "전쟁")\
#             .replace("19금의 코미디", "19금, 코미디")\
#             .replace("나가사키 켄지", "delete")\
#             .replace("잡지 연재", "delete")\
#             .replace("SF 학원 배틀", "SF, 학원, 배틀")\
#             .replace("드라마 스페이스 오페라", "드라마, 스페이스 오페라")\
#             .replace("오컬트 추리", "오컬트, 추리")\
#             .replace("공상과학 코미디", "공상과학, 코미디")\
#             .replace("누아르", "느와르")
#
#         work_book.save(xl)

# temp_data = []
#
# genre_list = ["코미디", "블랙 코미디", "아포칼립스", "SF", "판타지", "모험", "메카", "액션", "스릴러", "일상", "드라마", "백합",
#               "19금", "치유", "학원", "심리", "로맨스", "로맨틱 코미디", "가족", "드라마", "개그", "스포츠", "이세계", "미스터리",
#               "공포", "공상과학", "시대극", "초자연", "세카이"]
#
# remove_data = ["delete", "배틀액션", "성인식"]
#
# remove_index = []
#
# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"])):
#         tag_data = []
#         genre_data = []
#         for kk in sheet["B" + str(no)].value.split(", "):
#             if kk not in genre_list:
#                 tag_data.append(kk)
#             else:
#                 genre_data.append(kk)
#
#         sheet["B" + str(no)].value = ", ".join(genre_data)
#         sheet["C" + str(no)].value = ", ".join(tag_data)
#
#         work_book.save(xl)

# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"])):
#         for kk in sheet["B"+str(no)].value.split(", "):
#             if kk not in genre_data:
#                 temp_data.append(kk)
#         sheet["C" + str(no)].value = ", ".join(temp_data)

# total_data = list(set(temp_data))
# for index, nn in enumerate(total_data):
#     if index % 60 == 0:
#         print("\n\n")
#     print(nn)

# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"])):
#         for kk in sheet["B"+str(no)].value.split(", "):
#             if kk not in genre_list:
#                 temp_data.append(kk)
#
# total_data = list(set(temp_data))
#
# for index, nn in enumerate(total_data):
#     if index % 54 == 0:
#         print("\n\n")
#     print(nn)
#
# print(len(total_data))

# for xl in xlsx_list:
#     work_book = load_workbook(xl)
#     sheet = work_book['Sheet']
#     for no in range(2, len(sheet["A"])):
#         for kk in sheet["B"+str(no)].value.split(", "):
#             if kk in remove_data:
#                 remove_index.append(xl + " " + str(no))
#             elif kk == "" or kk is None:
#                 remove_index.append(xl + " " + str(no))
#
# print(remove_index)

