
from openpyxl.styles import Font, Alignment
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
import pandas as pd

from Settings.selenium_setting import *

# 크롤링 페이지
driver.get("https://namu.wiki/w/%EC%9D%BC%EB%B3%B8%20%EC%95%A0%EB%8B%88%EB%A9%94%EC%9D%B4%EC%85%98/%EB%AA%A9%EB%A1%9D"
           "/%E3%84%B1")

ani_data = {}
ani_title_list = []
xlsx_name = "ani_a"

# 애니 리스트 크롤링
ani_detail_list_xpath_base = "//*[@id='app']/div/div[2]/article/div[3]/div[2]/div/div/div/ul/li/div"
ani_detail_title = driver.find_elements_by_xpath(ani_detail_list_xpath_base + "/a")

for index, kk in enumerate(ani_detail_title):
    ani_data[kk.text] = kk.get_attribute("href")
    ani_title_list.append(kk.text)

driver.quit()

# 엑셀파일 쓰기
work_book = Workbook()

# 시트 생성
# work_book.create_sheet('애니메이션')

# 시트 입력
sheet = work_book.active
sheet.column_dimensions["A"].width = 55
sheet.column_dimensions["B"].width = 45
sheet.column_dimensions["C"].width = 45
sheet.column_dimensions["D"].width = 20
sheet.column_dimensions["E"].width = 45
sheet.column_dimensions["H"].width = 30

sheet['A1'] = "제목"
sheet['A1'].font = Font(name="나눔고딕", color="000000", bold=True)
sheet['A1'].alignment = Alignment(horizontal='center')
sheet['B1'] = "장르"
sheet['B1'].font = Font(name="나눔고딕", color="000000", bold=True)
sheet['B1'].alignment = Alignment(horizontal='center')
sheet['C1'] = "태그"
sheet['C1'].font = Font(name="나눔고딕", color="000000", bold=True)
sheet['C1'].alignment = Alignment(horizontal='center')
sheet['D1'] = "제작년도"
sheet['D1'].font = Font(name="나눔고딕", color="000000", bold=True)
sheet['D1'].alignment = Alignment(horizontal='center')
sheet['E1'] = "대표이미지"
sheet['E1'].font = Font(name="나눔고딕", color="000000", bold=True)
sheet['E1'].alignment = Alignment(horizontal='center')
sheet['H1'] = "링크"
sheet['H1'].font = Font(name="나눔고딕", color="000000", bold=True)
sheet['H1'].alignment = Alignment(horizontal='center')


for kk in range(0, len(ani_data)):
    sheet['A' + str(kk + 2)] = ani_title_list[kk]
    sheet['A' + str(kk + 2)].font = Font(name="나눔고딕", color="000000")
    sheet['H' + str(kk + 2)] = ani_data[ani_title_list[kk]]
    # sheet['H' + str(kk + 2)].value = '=HYPERLINK("{}", "{}")'.format(ani_data[ani_title_list[kk - 1]], "링크")
    sheet['H' + str(kk + 2)].font = Font(name="나눔고딕", color="50BCDF")

work_book.save(xlsx_name + ".xlsx")


