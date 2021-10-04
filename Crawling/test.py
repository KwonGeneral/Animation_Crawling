import datetime
import urllib.request

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.common.exceptions import StaleElementReferenceException, InvalidArgumentException, TimeoutException, \
    NoSuchElementException, WebDriverException
import time
import gdown

google_path = 'https://drive.google.com/uc?id='
file_id = '17KP7M-U7YSnqIp93QjnKiFH-pIUemL9i'
output_name = 'sample.xlsx'
gdown.download(google_path+file_id, output_name, quiet=False)

# chd = 'C:/dev_files/chd/chd.exe'
# options = webdriver.ChromeOptions()
# options.add_argument("headless")
# options.add_argument("window-size=1920x1080")
# options.add_argument("disable-gpu")
# options.add_argument("User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, "
#                      "like Gecko) Chrome/93.0.4577.82")
# options.add_argument("lang=ko_KR")
# driver = webdriver.Chrome(chd, options=options)
#
# load_book = load_workbook("ani_detail_link.xlsx")
# load_sheet = load_book['Sheet']
#
# work_book = load_workbook("ani_detail.xlsx")
# sheet = work_book["Sheet"]
#
# content_xpath = "//*[@id='animeContents']"
#
# for no in range(2, len(load_sheet["A"]) + 1):
#     try:
#         # 크롤링 주소 가져오기
#         driver.get(load_sheet["B" + str(no)].value)
#         time.sleep(5)
#
#         # XPATH 찾기
#         content_tag = driver.find_elements_by_xpath(content_xpath)
#
#         # 줄거리 크롤링
#         if len(content_tag) > 0:
#             for kk in content_tag:
#                 if str(kk.text).replace(" ", "") == "-":
#                     sheet["F" + str(no)].value = ""
#                     print("[ " + str(no) + " ] 줄거리 X : ", sheet["A" + str(no)].value)
#                 else:
#                     sheet["F" + str(no)].value = kk.text
#                     print("[ " + str(no) + " ] 줄거리 O : ", sheet["A" + str(no)].value)
#         else:
#             sheet["F" + str(no)].value = ""
#             print("[ " + str(no) + " ] 줄거리 X : ", sheet["A" + str(no)].value)
#
#     except StaleElementReferenceException:
#         print("StaleElementReferenceException")
#     except InvalidArgumentException:
#         print("InvalidArgumentException")
#     except NoSuchElementException:
#         print("NoSuchElementException")
#     except TimeoutException:
#         print("TimeoutException")
#         driver.quit()
#
#         time.sleep(5)
#         options = webdriver.ChromeOptions()
#         options.add_argument("headless")
#         options.add_argument("window-size=1920x1080")
#         options.add_argument("disable-gpu")
#         options.add_argument("User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, "
#                              "like Gecko) Chrome/93.0.4577.82")
#         options.add_argument("lang=ko_KR")
#         driver = webdriver.Chrome(chd, options=options)
#
#     except WebDriverException:
#         print("WebDriverException")
#     except IndexError:
#         print("IndexError")
#     except:
#         print("Error 발생!")
#
# driver.quit()
#
# work_book.save("ani_detail.xlsx")





# from openpyxl import load_workbook
# 
# a_xl = "ani_detail.xlsx"
# a_work_book = load_workbook(a_xl)
# a_sheet = a_work_book['Sheet']
# 
# b_xl = "ani_detail_link.xlsx"
# b_work_book = load_workbook(b_xl)
# b_sheet = b_work_book['Sheet']
# 
# a_list = []
# b_list = []
# 
# for no in range(2, len(a_sheet["A"])):
#     a_list.append(a_sheet["A"+str(no)].value)
# 
# for no in range(2, len(b_sheet["A"])):
#     b_list.append(b_sheet["A"+str(no)].value)
# 
# for bbb in b_list:
#     if bbb not in a_list:
#         print(bbb)

# update_file = open("update_log.txt", 'a')  # r : read, w : write, a : append
# update_list = update_file.read().split("\n")
# print(update_list)
# update_file.close()
# update_file.write("\n하하핳")

# xl = "ani_detail.xlsx"
# work_book = load_workbook(xl)
# sheet = work_book['Sheet']
# genre_list = ["스포츠", "추리", "코미디", "성전환", "먼치킨", "모험", "SF", "하렘", "드라마", "이세계",
#               "일상", "학원", "로맨스", "미스테리", "치유", "판타지", "메카", "스릴러", "19금",
#               "공포", "공상과학", "초자연", "음악", "아동", "액션", "BL", "백합"]
#
# for no in range(2, len(sheet["A"])):
#     temp_list = []
#     if sheet["B" + str(no)].value is not None:
#         for kk in sheet["B" + str(no)].value.split(", "):
#             if len(sheet["B" + str(no)].value.split(", ")) == 1 and kk not in genre_list:
#                 temp_str = "[ " + str(no) + " ] " + sheet["A" + str(no)].value + "[ " + \
#                            sheet["B" + str(no)].value + " ]"
#                 temp_list.append(temp_str)
#     for a in temp_list:
#         print(a)

# now_time = datetime.datetime.now()
# front_time = str(now_time.year) + str(now_time.month) + str(now_time.day)
# back_time = str(now_time.hour) + str(now_time.minute) + str(now_time.second) + str(now_time.microsecond)
# image_file_name = front_time + back_time + ".jpg"
# print(image_file_name)

# 맨 우측 콤마(, ) 제거
# xl = "ani_detail.xlsx"
# work_book = load_workbook(xl)
# sheet = work_book['Sheet']
# for no in range(2, len(sheet["A"])):
#     if sheet["B"+str(no)].value is not None:
#         sheet["B"+str(no)].value = sheet["B"+str(no)].value.rstrip(", ")
#
# work_book.save(xl)


# genre_list = ['로리', '역사', '쇼타', '소꿉친구', '마왕', '자동차', '버추얼 리얼리티', '성우', '호러', '퇴마',
#               '괴물', '천사', '밀리터리', '코미디', '성년', '영웅', '아이돌', '멘붕', '패러디', '공상과학',
#               '심리', '메카', '정령', '동물', '슈퍼파워', '슈팅', '범죄', '게임', '변신', '시뮬레이션', '공포',
#               '귀신', '경찰', '여성', '전투', '드래곤', '연애', '액션', '성전환', '격투', '전쟁', '유령',
#               '미스테리', 'SF', '사무라이', '추리', 'BL', '학원', '모험', '스릴러', '재판', '메카닉', '초자연',
#               '마법소녀', '신', '배틀', '로맨스', '집사', '좀비', '요괴', '먼치킨', '하렘', '일상', '메이드',
#               '우주', '드라마', '악마', '판타지', '음악', '19금', '마녀', '소녀', '이세계', '부활동', '시대',
#               '군', '요리', '백합', '리듬', '디멘시아', '능력', '아동', '소년', '치유', 'None', '스포츠',
#               '마법', '카페', '닌자', '미스터리', '뱀파이어']
